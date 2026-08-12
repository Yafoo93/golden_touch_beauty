import csv
from io import BytesIO, StringIO
from pathlib import Path

from django.conf import settings
from django.http import HttpResponse
from django.utils import timezone
from openpyxl import Workbook
from openpyxl.drawing.image import Image as ExcelImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import mm
from reportlab.platypus import KeepTogether, Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from branches.models import Branch


GOLD = colors.HexColor("#D8A622")
BLACK = colors.HexColor("#080808")
CREAM = colors.HexColor("#F7F1E5")
GREY = colors.HexColor("#555555")

REPORTS = {
    "sales": {
        "title": "Online Orders and POS Sales Report",
        "summary": ["total_revenue", "online_revenue", "pos_revenue", "transaction_count", "online_count", "pos_count", "average_sale"],
        "tables": [("Transactions", "transactions", ["occurred_at", "reference", "source", "branch_name", "customer_name", "status", "amount"])],
    },
    "bookings": {
        "title": "Appointments Report",
        "summary": ["booking_count", "active_count", "completed_count", "cancelled_count", "cancellation_rate", "no_show_count", "no_show_rate", "booked_value"],
        "tables": [("Appointments", "bookings", ["preferred_start", "reference", "branch_name", "customer_name", "service_names", "status", "payment_status", "amount"])],
    },
    "products": {
        "title": "Product Performance Report",
        "summary": ["product_count", "units_sold", "revenue", "cost_of_goods", "gross_profit", "gross_margin_percent", "low_stock_count", "out_of_stock_count"],
        "tables": [("Products", "products", ["name", "variant", "sku", "units_sold", "revenue", "cost_of_goods", "gross_profit", "gross_margin_percent", "stock_available", "stock_state"])],
    },
    "services": {
        "title": "Service Performance Report",
        "summary": ["service_count", "distinct_services", "revenue", "average_value", "completed_bookings", "duration_minutes"],
        "tables": [("Services", "performance", ["name", "booking_count", "pos_count", "service_count", "booking_revenue", "pos_revenue", "revenue", "average_value", "completed_count"])],
    },
    "inventory": {
        "title": "Stock Levels and Movements Report",
        "summary": ["inventory_count", "quantity_on_hand", "quantity_reserved", "quantity_available", "cost_value", "retail_value", "low_stock_count", "out_of_stock_count", "movement_count", "on_hand_change"],
        "tables": [
            ("Current Stock", "inventory", ["branch_name", "product_name", "variant_name", "sku", "quantity_on_hand", "quantity_reserved", "quantity_available", "reorder_level", "stock_state", "cost_value", "retail_value"]),
            ("Movement Ledger", "movements", ["occurred_at", "branch_name", "product_name", "variant_name", "sku", "type", "on_hand_change", "reserved_change", "on_hand_after", "reserved_after", "reference_id", "note", "performed_by"]),
        ],
    },
    "payments": {
        "title": "Payments Reconciliation Report",
        "summary": ["payment_count", "successful_count", "successful_amount", "pending_count", "failed_count", "refunded_count", "refunded_amount", "net_collected"],
        "tables": [
            ("Payment-method Totals", "by_method", ["method", "attempted_count", "successful_count", "collected_amount", "refunded_amount", "net_collected", "online_amount", "pos_amount"]),
            ("Payment Activity", "payments", ["occurred_at", "reference", "source", "source_reference", "branch_name", "customer_name", "provider", "method", "status", "amount"]),
        ],
    },
    "branches": {
        "title": "Branch Comparison Report",
        "disclaimer": "Estimated operating result is not net profit. Service consumables, labour, commissions, delivery costs, rent, utilities, taxes, and other operating expenses are not yet captured.",
        "summary": ["branch_count", "total_sales", "booking_count", "payments_collected", "product_revenue", "product_gross_profit", "service_revenue", "estimated_operating_result", "stock_available"],
        "tables": [("Branch Performance", "performance", ["branch_name", "total_sales", "sales_share_percent", "online_sales", "pos_sales", "booking_count", "completed_bookings", "cancellation_rate", "no_show_rate", "product_revenue", "product_gross_profit", "service_revenue", "estimated_operating_result", "payments_collected", "stock_available", "low_stock_count", "out_of_stock_count"])],
    },
}


def _label(value):
    return str(value).replace("_", " ").strip().title()


def _display(value):
    if value is None or value == "":
        return "—"
    if isinstance(value, (list, tuple)):
        return ", ".join(map(str, value)) or "—"
    if isinstance(value, bool):
        return "Yes" if value else "No"
    return str(value)


def _spreadsheet_display(value):
    """Render safely without allowing user text to become a spreadsheet formula."""
    rendered = _display(value)
    if isinstance(value, str) and rendered.startswith(("=", "+", "-", "@")):
        return f"'{rendered}"
    return rendered


def _logo_path():
    candidates = [
        Path(settings.BASE_DIR).parent / "frontend" / "public" / "images" / "logo.png",
        Path(settings.BASE_DIR).parent / "frontend" / "public" / "images" / "logo2.png",
    ]
    return next((path for path in candidates if path.exists()), None)


def _branch_lines():
    lines = []
    for branch in Branch.objects.filter(is_active=True).order_by("name"):
        contacts = " / ".join(filter(None, [branch.telephone_number, branch.secondary_telephone_number]))
        hours = f"{branch.opening_time:%H:%M}–{branch.closing_time:%H:%M}"
        lines.append(f"{branch.name}: {branch.address} | {contacts} | {hours}")
    return lines


def _metadata(report_name, payload, user):
    filters = payload.get("filters", {})
    selected = ", ".join(f"{_label(key)}: {_display(value)}" for key, value in filters.items() if value not in (None, "", "all"))
    return {
        "report": REPORTS[report_name]["title"],
        "period": f"{filters.get('date_from', 'All dates')} to {filters.get('date_to', 'All dates')}",
        "filters": selected or "All permitted data",
        "generated": timezone.localtime().strftime("%d %B %Y, %H:%M %Z"),
        "generated_by": user.full_name or user.email,
    }


def export_csv(report_name, payload, user):
    config = REPORTS[report_name]
    output = StringIO(newline="")
    writer = csv.writer(output)
    metadata = _metadata(report_name, payload, user)
    writer.writerow([_spreadsheet_display(metadata["report"])])
    writer.writerow(["Period", _spreadsheet_display(metadata["period"])])
    writer.writerow(["Filters", _spreadsheet_display(metadata["filters"])])
    writer.writerow(["Generated", _spreadsheet_display(metadata["generated"])])
    writer.writerow(["Generated by", _spreadsheet_display(metadata["generated_by"])])
    if config.get("disclaimer"):
        writer.writerow(["Important limitation", _spreadsheet_display(config["disclaimer"])])
    writer.writerow([])
    writer.writerow(["Summary"])
    for key in config["summary"]:
        writer.writerow([_label(key), _spreadsheet_display(payload.get("summary", {}).get(key))])
    for title, data_key, columns in config["tables"]:
        writer.writerow([])
        writer.writerow([title])
        writer.writerow([_label(column) for column in columns])
        for row in payload.get(data_key, []):
            writer.writerow([_spreadsheet_display(row.get(column)) for column in columns])
    response = HttpResponse(output.getvalue().encode("utf-8-sig"), content_type="text/csv; charset=utf-8")
    response["Content-Disposition"] = f'attachment; filename="golden-touch-{report_name}-report.csv"'
    return response


def export_excel(report_name, payload, user):
    config = REPORTS[report_name]
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.sheet_view.showGridLines = False
    metadata = _metadata(report_name, payload, user)
    summary_sheet.merge_cells("A1:D1")
    summary_sheet["A1"] = metadata["report"]
    summary_sheet["A1"].font = Font(size=18, bold=True, color="D8A622")
    summary_sheet["A2"] = "Golden Touch Beauty Centre"
    summary_sheet["A2"].font = Font(size=12, bold=True)
    for index, key in enumerate(("period", "filters", "generated", "generated_by"), start=4):
        summary_sheet.cell(index, 1, _label(key)).font = Font(bold=True)
        summary_sheet.cell(index, 2, _spreadsheet_display(metadata[key]))
    if config.get("disclaimer"):
        summary_sheet["A8"] = "Important limitation"
        summary_sheet["A8"].font = Font(bold=True, color="9C6A00")
        summary_sheet["B8"] = _spreadsheet_display(config["disclaimer"])
        summary_sheet["B8"].alignment = Alignment(wrap_text=True)
    row_index = 9
    summary_sheet.cell(row_index, 1, "Metric").font = Font(bold=True, color="FFFFFF")
    summary_sheet.cell(row_index, 2, "Value").font = Font(bold=True, color="FFFFFF")
    for cell in summary_sheet[row_index][:2]:
        cell.fill = PatternFill("solid", fgColor="080808")
    for key in config["summary"]:
        row_index += 1
        summary_sheet.cell(row_index, 1, _label(key))
        summary_sheet.cell(row_index, 2, _spreadsheet_display(payload.get("summary", {}).get(key)))
    summary_sheet.column_dimensions["A"].width = 28
    summary_sheet.column_dimensions["B"].width = 65
    logo = _logo_path()
    if logo:
        image = ExcelImage(str(logo))
        image.width, image.height = 64, 64
        summary_sheet.add_image(image, "D2")

    for title, data_key, columns in config["tables"]:
        sheet = workbook.create_sheet(title[:31])
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        for column_index, column in enumerate(columns, start=1):
            cell = sheet.cell(1, column_index, _label(column))
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor="080808")
            cell.alignment = Alignment(horizontal="center")
        for row_index, row in enumerate(payload.get(data_key, []), start=2):
            for column_index, column in enumerate(columns, start=1):
                sheet.cell(row_index, column_index, _spreadsheet_display(row.get(column)))
        sheet.auto_filter.ref = sheet.dimensions
        for column_index, column in enumerate(columns, start=1):
            values = [_label(column)] + [_spreadsheet_display(row.get(column)) for row in payload.get(data_key, [])[:200]]
            sheet.column_dimensions[get_column_letter(column_index)].width = min(max(map(len, values)) + 2, 40)

    stream = BytesIO()
    workbook.save(stream)
    response = HttpResponse(stream.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = f'attachment; filename="golden-touch-{report_name}-report.xlsx"'
    return response


def export_pdf(report_name, payload, user):
    config = REPORTS[report_name]
    metadata = _metadata(report_name, payload, user)
    stream = BytesIO()
    document = SimpleDocTemplate(
        stream, pagesize=landscape(A4), leftMargin=14 * mm, rightMargin=14 * mm,
        topMargin=43 * mm, bottomMargin=18 * mm, title=metadata["report"],
        author="Golden Touch Beauty Centre",
    )
    styles = getSampleStyleSheet()
    styles.add(ParagraphStyle(name="ReportTitle", parent=styles["Title"], fontName="Helvetica-Bold", fontSize=18, textColor=BLACK, spaceAfter=5 * mm))
    styles.add(ParagraphStyle(name="Section", parent=styles["Heading2"], fontName="Helvetica-Bold", fontSize=12, textColor=BLACK, spaceBefore=4 * mm, spaceAfter=2 * mm))
    styles.add(ParagraphStyle(name="Meta", parent=styles["BodyText"], fontSize=8, leading=11, textColor=GREY))
    styles.add(ParagraphStyle(name="Cell", parent=styles["BodyText"], fontSize=6.5, leading=8))
    styles.add(ParagraphStyle(name="CellHead", parent=styles["BodyText"], fontName="Helvetica-Bold", fontSize=6.5, leading=8, textColor=colors.white, alignment=TA_CENTER))

    branches = _branch_lines()
    logo = _logo_path()

    def letterhead(canvas, doc):
        canvas.saveState()
        width, height = landscape(A4)
        canvas.setFillColor(BLACK)
        canvas.rect(0, height - 32 * mm, width, 32 * mm, fill=1, stroke=0)
        if logo:
            try:
                canvas.drawImage(str(logo), 14 * mm, height - 27 * mm, 20 * mm, 20 * mm, preserveAspectRatio=True, mask="auto")
            except Exception:
                pass
        canvas.setFillColor(GOLD)
        canvas.setFont("Helvetica-Bold", 17)
        canvas.drawString(38 * mm, height - 13 * mm, "Golden Touch Beauty Centre")
        canvas.setFillColor(colors.white)
        canvas.setFont("Helvetica", 7.5)
        y = height - 18 * mm
        for line in branches[:3]:
            canvas.drawString(38 * mm, y, line[:155])
            y -= 4 * mm
        canvas.setStrokeColor(GOLD)
        canvas.setLineWidth(1.5)
        canvas.line(0, height - 32 * mm, width, height - 32 * mm)
        canvas.setFont("Helvetica", 7)
        canvas.setFillColor(GREY)
        canvas.drawString(14 * mm, 9 * mm, "Confidential management report · Golden Touch Beauty Centre")
        canvas.drawRightString(width - 14 * mm, 9 * mm, f"Page {doc.page}")
        canvas.restoreState()

    story = [Paragraph(metadata["report"], styles["ReportTitle"])]
    story.append(Paragraph(f"<b>Reporting period:</b> {metadata['period']}<br/><b>Selected filters:</b> {metadata['filters']}<br/><b>Generated:</b> {metadata['generated']} by {metadata['generated_by']}", styles["Meta"]))
    if config.get("disclaimer"):
        story.append(Spacer(1, 2 * mm))
        story.append(Paragraph(f"<b>Important limitation:</b> {config['disclaimer']}", styles["Meta"]))
    story.append(Spacer(1, 4 * mm))
    summary_rows = [[Paragraph("Executive Summary", styles["CellHead"]), Paragraph("Value", styles["CellHead"])]]
    for key in config["summary"]:
        summary_rows.append([Paragraph(_label(key), styles["Cell"]), Paragraph(_display(payload.get("summary", {}).get(key)), styles["Cell"])])
    summary_table = Table(summary_rows, colWidths=[70 * mm, 45 * mm], repeatRows=1)
    summary_table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), BLACK), ("GRID", (0, 0), (-1, -1), 0.35, colors.HexColor("#CFC7B8")), ("BACKGROUND", (0, 1), (-1, -1), CREAM), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5), ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4)]))
    story.append(KeepTogether([summary_table]))

    usable_width = landscape(A4)[0] - 28 * mm
    for title, data_key, columns in config["tables"]:
        story.append(Paragraph(title, styles["Section"]))
        data = payload.get(data_key, [])
        if not data:
            story.append(Paragraph("No records matched the selected filters.", styles["Meta"]))
            continue
        table_rows = [[Paragraph(_label(column), styles["CellHead"]) for column in columns]]
        for row in data[:500]:
            table_rows.append([Paragraph(_display(row.get(column)).replace("&", "&amp;").replace("<", "&lt;"), styles["Cell"]) for column in columns])
        table = Table(table_rows, colWidths=[usable_width / len(columns)] * len(columns), repeatRows=1)
        table.setStyle(TableStyle([("BACKGROUND", (0, 0), (-1, 0), BLACK), ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#D8D2C6")), ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, CREAM]), ("VALIGN", (0, 0), (-1, -1), "TOP"), ("LEFTPADDING", (0, 0), (-1, -1), 2.5), ("RIGHTPADDING", (0, 0), (-1, -1), 2.5), ("TOPPADDING", (0, 0), (-1, -1), 3), ("BOTTOMPADDING", (0, 0), (-1, -1), 3)]))
        story.append(table)
        if len(data) > 500:
            story.append(Paragraph(f"PDF limited to the first 500 of {len(data)} records. Use Excel or CSV for the complete dataset.", styles["Meta"]))
    document.build(story, onFirstPage=letterhead, onLaterPages=letterhead)
    response = HttpResponse(stream.getvalue(), content_type="application/pdf")
    response["Content-Disposition"] = f'attachment; filename="golden-touch-{report_name}-report.pdf"'
    return response
