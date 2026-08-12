from datetime import datetime, time
from decimal import Decimal
from io import BytesIO

from django.test import TestCase
from django.urls import reverse
from django.utils import timezone
from rest_framework import status
from openpyxl import load_workbook

from accounts.models import User
from branches.models import Branch, BranchStaffAssignment
from bookings.models import Booking, BookingServiceItem
from inventory.models import BranchInventory, StockMovement
from orders.models import Order, OrderItem
from payments.models import Payment
from pos.models import POSPaymentEntry, POSSale, POSSaleLine
from products.models import Product, ProductCategory, ProductVariant
from services.models import Service, ServiceBranchAvailability, ServiceCategory


class ManagementSalesReportTests(TestCase):
    def setUp(self):
        self.makola = Branch.objects.create(
            name="Report Makola", code="REPORT-MAK", address="Accra",
            telephone_number="+233200001001", opening_days=["monday"],
            opening_time=time(7, 30), closing_time=time(17),
        )
        self.tse_addo = Branch.objects.create(
            name="Report Tse Addo", code="REPORT-TSE", address="Accra",
            telephone_number="+233200001002", opening_days=["monday"],
            opening_time=time(7, 30), closing_time=time(19),
        )
        self.owner = User.objects.create_superuser(
            email="report-owner@example.com", phone_number="+233200001003",
            full_name="Report Owner", password="OwnerPass123!",
        )
        self.manager = User.objects.create_user(
            email="report-manager@example.com", phone_number="+233200001004",
            full_name="Report Manager", password="ManagerPass123!", is_staff=True,
        )
        self.cashier = User.objects.create_user(
            email="report-cashier@example.com", phone_number="+233200001005",
            full_name="Report Cashier", password="CashierPass123!", is_staff=True,
        )
        self.customer = User.objects.create_user(
            email="report-customer@example.com", phone_number="+233200001006",
            full_name="Report Customer", password="CustomerPass123!",
        )
        BranchStaffAssignment.objects.create(
            branch=self.makola, staff=self.manager,
            roles=[BranchStaffAssignment.Role.MANAGER], assigned_by=self.owner,
        )
        BranchStaffAssignment.objects.create(
            branch=self.makola, staff=self.cashier,
            roles=[BranchStaffAssignment.Role.CASHIER], assigned_by=self.owner,
        )
        now = timezone.now()
        self.order = Order.objects.create(
            branch=self.makola, customer=self.customer, status=Order.Status.PAID,
            payment_status="paid", subtotal="100.00", total_amount="100.00",
            recipient_name=self.customer.full_name, paid_at=now,
        )
        Payment.objects.create(
            branch=self.makola, customer=self.customer, order=self.order,
            provider="paystack", method="mobile_money", status=Payment.Status.SUCCEEDED,
            amount="100.00", paid_at=now,
        )
        Payment.objects.create(
            branch=self.makola, customer=self.customer, order=self.order,
            provider="paystack", method="card", status=Payment.Status.FAILED,
            amount="20.00",
        )
        category = ProductCategory.objects.create(name="Report products", slug="report-products")
        product = Product.objects.create(category=category, name="Report Serum", slug="report-serum", description="Test", is_active=True, is_published=True)
        self.variant = ProductVariant.objects.create(product=product, name="Standard", sku="REPORT-SERUM", selling_price="25.00", cost_price="10.00")
        OrderItem.objects.create(order=self.order, product_variant=self.variant, product_name=product.name, product_slug=product.slug, variant_name=self.variant.name, sku=self.variant.sku, unit_price="25.00", unit_cost="10.00", quantity=4, line_total="100.00", line_cost="40.00")
        self.inventory = BranchInventory.objects.create(branch=self.makola, product_variant=self.variant, quantity_on_hand=5, quantity_reserved=1, reorder_level=5)
        StockMovement.objects.create(inventory=self.inventory, movement_type=StockMovement.MovementType.OPENING, quantity_on_hand_change=5, quantity_reserved_change=1, quantity_on_hand_after=5, quantity_reserved_after=1, note="Report opening stock", performed_by=self.owner)
        self.pos_sale = POSSale.objects.create(
            branch=self.makola, cashier=self.cashier, customer=self.customer,
            status=POSSale.Status.COMPLETED, payment_status="paid",
            total_amount="50.00", item_count=1, completed_at=now,
        )
        POSPaymentEntry.objects.create(sale=self.pos_sale, method="cash", amount="50.00")
        POSSaleLine.objects.create(sale=self.pos_sale, item_type=POSSaleLine.ItemType.PRODUCT, item_reference=str(self.variant.pk), name=product.name, option_name=self.variant.name, sku=self.variant.sku, quantity=2, unit_price="25.00", unit_cost="10.00", line_total="50.00", line_cost="20.00")
        POSSale.objects.create(
            branch=self.tse_addo, cashier=self.owner, status=POSSale.Status.COMPLETED,
            payment_status="paid", total_amount="999.00", item_count=1, completed_at=now,
        )
        self.makola_booking = Booking.objects.create(
            branch=self.makola, customer=self.customer, status=Booking.Status.CONFIRMED,
            source=Booking.Source.WEBSITE, preferred_start=now, total_amount="200.00",
            total_duration_minutes=90, payment_status="paid",
        )
        completed_booking = Booking.objects.create(
            branch=self.makola, customer=self.customer, status=Booking.Status.COMPLETED,
            source=Booking.Source.WALK_IN, preferred_start=now, total_amount="100.00",
            total_duration_minutes=60, payment_status="paid",
        )
        service_category = ServiceCategory.objects.create(name="Report services", slug="report-services")
        self.service = Service.objects.create(category=service_category, name="Report Facial", slug="report-facial", short_description="Test", description="Test", price="100.00", duration_minutes=60, is_active=True, is_published=True)
        ServiceBranchAvailability.objects.create(service=self.service, branch=self.makola, is_available=True)
        BookingServiceItem.objects.create(booking=self.makola_booking, service=self.service, service_name=self.service.name, unit_price="100.00", duration_minutes=60)
        BookingServiceItem.objects.create(booking=completed_booking, service=self.service, service_name=self.service.name, unit_price="100.00", duration_minutes=60)
        POSSaleLine.objects.create(sale=self.pos_sale, item_type=POSSaleLine.ItemType.SERVICE, item_reference=str(self.service.pk), name=self.service.name, quantity=1, unit_price="80.00", line_total="80.00")
        Booking.objects.create(
            branch=self.tse_addo, customer=self.customer, status=Booking.Status.CONFIRMED,
            source=Booking.Source.WHATSAPP, preferred_start=now, total_amount="999.00",
            total_duration_minutes=120,
        )

    def test_assigned_manager_sees_online_and_pos_sales_for_their_branch(self):
        self.client.force_login(self.manager)
        response = self.client.get(reverse("reports:sales"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.json()["summary"], {
            "total_revenue": "150.00", "online_revenue": "100.00",
            "pos_revenue": "50.00", "transaction_count": 2,
            "online_count": 1, "pos_count": 1,
            "online_share_percent": "66.67", "pos_share_percent": "33.33",
            "online_average_sale": "100.00", "pos_average_sale": "50.00",
            "average_sale": "75.00",
        })
        self.assertEqual({item["source"] for item in response.json()["transactions"]}, {"online", "pos"})
        self.assertEqual({item["method"] for item in response.json()["payment_methods"]}, {"cash", "mobile_money"})
        self.assertEqual([branch["id"] for branch in response.json()["branches"]], [str(self.makola.pk)])

    def test_source_filter_and_invalid_branch_are_enforced(self):
        self.client.force_login(self.manager)
        filtered = self.client.get(reverse("reports:sales"), {"source": "pos"})
        denied = self.client.get(reverse("reports:sales"), {"branch": str(self.tse_addo.pk)})

        self.assertEqual(filtered.status_code, status.HTTP_200_OK)
        self.assertEqual(filtered.json()["summary"]["total_revenue"], "50.00")
        self.assertEqual(filtered.json()["summary"]["online_revenue"], "0.00")
        self.assertEqual(filtered.json()["summary"]["online_count"], 0)
        self.assertEqual(filtered.json()["summary"]["pos_count"], 1)
        self.assertEqual(filtered.json()["summary"]["pos_share_percent"], "100.00")
        self.assertEqual(denied.status_code, status.HTTP_400_BAD_REQUEST)

    def test_sales_report_supports_daily_weekly_and_monthly_intervals(self):
        recent = timezone.make_aware(datetime(2026, 3, 12, 10, 0))
        older = timezone.make_aware(datetime(2026, 3, 3, 10, 0))
        self.order.paid_at = recent
        self.order.save(update_fields=["paid_at"])
        POSSale.objects.filter(pk=self.pos_sale.pk).update(completed_at=recent)
        Order.objects.create(
            branch=self.makola, customer=self.customer, status=Order.Status.PAID,
            payment_status="paid", subtotal="30.00", total_amount="30.00",
            recipient_name=self.customer.full_name, paid_at=older,
        )
        self.client.force_login(self.manager)
        date_from = "2026-03-01"
        date_to = "2026-03-31"
        daily = self.client.get(reverse("reports:sales"), {"date_from": date_from, "date_to": date_to, "interval": "daily"})
        weekly = self.client.get(reverse("reports:sales"), {"date_from": date_from, "date_to": date_to, "interval": "weekly"})
        monthly = self.client.get(reverse("reports:sales"), {"date_from": date_from, "date_to": date_to, "interval": "monthly"})

        self.assertEqual(len(daily.json()["trend"]), 2)
        self.assertEqual(len(weekly.json()["trend"]), 2)
        self.assertEqual(len(monthly.json()["trend"]), 1)
        for response in (daily, weekly, monthly):
            self.assertEqual(sum(Decimal(item["total"]) for item in response.json()["trend"]), Decimal("180.00"))

    def test_cashier_cannot_access_management_sales_report(self):
        self.client.force_login(self.cashier)
        response = self.client.get(reverse("reports:sales"))
        self.assertEqual(response.status_code, status.HTTP_403_FORBIDDEN)

    def test_booking_report_is_aggregated_and_branch_scoped(self):
        self.client.force_login(self.manager)
        response = self.client.get(reverse("reports:bookings"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.json()["summary"], {
            "booking_count": 2, "active_count": 2, "completed_count": 1,
            "cancelled_count": 0, "cancellation_rate": "0.00",
            "no_show_count": 0, "no_show_rate": "0.00", "rejected_count": 0,
            "booked_value": "300.00",
            "average_value": "150.00", "total_duration_minutes": 150,
        })
        self.assertEqual({row["source"] for row in response.json()["by_source"]}, {"website", "walk_in"})
        self.assertEqual([branch["id"] for branch in response.json()["branches"]], [str(self.makola.pk)])

    def test_booking_report_separates_cancellations_and_no_shows(self):
        now = timezone.now()
        Booking.objects.create(
            branch=self.makola, customer=self.customer,
            status=Booking.Status.CANCELLED, source=Booking.Source.PHONE,
            preferred_start=now, total_amount="50.00", total_duration_minutes=60,
        )
        Booking.objects.create(
            branch=self.makola, customer=self.customer,
            status=Booking.Status.NO_SHOW, source=Booking.Source.PHONE,
            preferred_start=now, total_amount="50.00", total_duration_minutes=60,
        )
        self.client.force_login(self.manager)
        response = self.client.get(reverse("reports:bookings"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        summary = response.json()["summary"]
        self.assertEqual(summary["booking_count"], 4)
        self.assertEqual(summary["cancelled_count"], 1)
        self.assertEqual(summary["cancellation_rate"], "25.00")
        self.assertEqual(summary["no_show_count"], 1)
        self.assertEqual(summary["no_show_rate"], "25.00")
        self.assertEqual(sum(day["cancelled_count"] for day in response.json()["daily"]), 1)
        self.assertEqual(sum(day["no_show_count"] for day in response.json()["daily"]), 1)

    def test_booking_report_filters_and_permissions_are_enforced(self):
        self.client.force_login(self.manager)
        filtered = self.client.get(reverse("reports:bookings"), {"status": Booking.Status.COMPLETED})
        denied_branch = self.client.get(reverse("reports:bookings"), {"branch": str(self.tse_addo.pk)})
        self.client.force_login(self.cashier)
        denied_role = self.client.get(reverse("reports:bookings"))

        self.assertEqual(filtered.status_code, status.HTTP_200_OK)
        self.assertEqual(filtered.json()["summary"]["booking_count"], 1)
        self.assertEqual(denied_branch.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(denied_role.status_code, status.HTTP_403_FORBIDDEN)

    def test_product_report_combines_online_pos_and_stock(self):
        self.variant.cost_price = Decimal("999.00")
        self.variant.save(update_fields=["cost_price"])
        self.client.force_login(self.manager)
        response = self.client.get(reverse("reports:products"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.json()["summary"], {
            "product_count": 1, "units_sold": 6, "revenue": "150.00",
            "cost_of_goods": "60.00", "gross_profit": "90.00",
            "gross_margin_percent": "60.00", "average_unit_revenue": "25.00", "low_stock_count": 1,
            "out_of_stock_count": 0,
        })
        self.assertEqual(response.json()["products"][0]["stock_available"], 4)
        self.assertEqual(response.json()["products"][0]["online_units"], 4)
        self.assertEqual(response.json()["products"][0]["pos_units"], 2)
        self.assertEqual(response.json()["products"][0]["gross_profit"], "90.00")
        self.assertEqual(response.json()["best_selling_products"], [{
            "rank": 1, "name": "Report Serum", "variant": "Standard",
            "sku": "REPORT-SERUM", "units_sold": 6, "revenue": "150.00",
        }])

    def test_product_report_filters_and_permissions_are_enforced(self):
        self.client.force_login(self.manager)
        filtered = self.client.get(reverse("reports:products"), {"source": "pos", "stock": "low"})
        denied_branch = self.client.get(reverse("reports:products"), {"branch": str(self.tse_addo.pk)})
        self.client.force_login(self.cashier)
        denied_role = self.client.get(reverse("reports:products"))

        self.assertEqual(filtered.status_code, status.HTTP_200_OK)
        self.assertEqual(filtered.json()["summary"]["revenue"], "50.00")
        self.assertEqual(denied_branch.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(denied_role.status_code, status.HTTP_403_FORBIDDEN)

    def test_service_report_combines_bookings_and_pos_services(self):
        unpaid_booking = Booking.objects.create(
            branch=self.makola, customer=self.customer,
            status=Booking.Status.CONFIRMED, source=Booking.Source.PHONE,
            preferred_start=timezone.now(), total_amount="100.00",
            total_duration_minutes=60, payment_status="pending",
        )
        BookingServiceItem.objects.create(
            booking=unpaid_booking, service=self.service,
            service_name=self.service.name, unit_price="100.00",
            duration_minutes=60,
        )
        self.client.force_login(self.manager)
        response = self.client.get(reverse("reports:services"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.json()["summary"], {
            "service_count": 4, "distinct_services": 1, "revenue": "280.00",
            "average_value": "70.00", "completed_bookings": 1,
            "duration_minutes": 180,
        })
        row = response.json()["performance"][0]
        self.assertEqual(row["booking_count"], 3)
        self.assertEqual(row["pos_count"], 1)
        self.assertEqual(row["booking_revenue"], "200.00")
        self.assertEqual(row["pos_revenue"], "80.00")
        self.assertEqual(response.json()["popular_services"], [{
            "rank": 1, "id": str(self.service.pk), "name": "Report Facial",
            "service_count": 4, "booking_count": 3, "pos_count": 1,
            "revenue": "280.00",
        }])

    def test_service_report_filters_and_permissions_are_enforced(self):
        self.client.force_login(self.manager)
        filtered = self.client.get(reverse("reports:services"), {"source": "booking", "status": Booking.Status.COMPLETED, "service": str(self.service.pk)})
        denied_branch = self.client.get(reverse("reports:services"), {"branch": str(self.tse_addo.pk)})
        self.client.force_login(self.cashier)
        denied_role = self.client.get(reverse("reports:services"))

        self.assertEqual(filtered.status_code, status.HTTP_200_OK)
        self.assertEqual(filtered.json()["summary"]["service_count"], 1)
        self.assertEqual(filtered.json()["summary"]["revenue"], "100.00")
        self.assertEqual(denied_branch.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(denied_role.status_code, status.HTTP_403_FORBIDDEN)

    def test_inventory_report_combines_position_valuation_and_movements(self):
        self.client.force_login(self.manager)
        response = self.client.get(reverse("reports:inventory"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.json()["summary"], {
            "inventory_count": 1, "quantity_on_hand": 5, "quantity_reserved": 1,
            "quantity_available": 4, "cost_value": "40.00", "retail_value": "100.00",
            "low_stock_count": 1, "out_of_stock_count": 0,
            "movement_count": 1, "on_hand_change": 5,
        })
        self.assertEqual(response.json()["inventory"][0]["stock_state"], "low")
        self.assertEqual(response.json()["movements_by_type"][0]["type"], StockMovement.MovementType.OPENING)
        self.assertEqual(len(response.json()["movements"]), 1)
        movement = response.json()["movements"][0]
        self.assertEqual(movement["sku"], "REPORT-SERUM")
        self.assertEqual(movement["on_hand_change"], 5)
        self.assertEqual(movement["on_hand_after"], 5)
        self.assertEqual(movement["performed_by"], "Report Owner")

    def test_inventory_report_filters_and_permissions_are_enforced(self):
        self.client.force_login(self.manager)
        filtered = self.client.get(reverse("reports:inventory"), {"stock": "low", "search": "SERUM"})
        denied_branch = self.client.get(reverse("reports:inventory"), {"branch": str(self.tse_addo.pk)})
        self.client.force_login(self.cashier)
        denied_role = self.client.get(reverse("reports:inventory"))

        self.assertEqual(filtered.status_code, status.HTTP_200_OK)
        self.assertEqual(filtered.json()["summary"]["inventory_count"], 1)
        self.assertEqual(denied_branch.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(denied_role.status_code, status.HTTP_403_FORBIDDEN)

    def test_payments_report_reconciles_online_and_pos_activity(self):
        self.client.force_login(self.manager)
        response = self.client.get(reverse("reports:payments"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.json()["summary"], {
            "payment_count": 3, "successful_count": 2,
            "successful_amount": "150.00", "pending_count": 0,
            "failed_count": 1, "refunded_count": 0,
            "refunded_amount": "0.00", "net_collected": "150.00",
        })
        self.assertEqual({row["source"] for row in response.json()["payments"]}, {"online", "pos"})
        self.assertEqual({row["method"] for row in response.json()["by_method"]}, {"card", "cash", "mobile_money"})
        methods = {row["method"]: row for row in response.json()["by_method"]}
        self.assertEqual(methods["mobile_money"], {
            "method": "mobile_money", "attempted_count": 1,
            "successful_count": 1, "collected_amount": "100.00",
            "refunded_amount": "0.00", "net_collected": "100.00",
            "online_amount": "100.00", "pos_amount": "0.00",
        })
        self.assertEqual(methods["cash"]["net_collected"], "50.00")
        self.assertEqual(methods["card"]["successful_count"], 0)
        self.assertEqual(methods["card"]["net_collected"], "0.00")

    def test_payments_report_filters_and_permissions_are_enforced(self):
        self.client.force_login(self.manager)
        filtered = self.client.get(reverse("reports:payments"), {"source": "online", "status": Payment.Status.SUCCEEDED, "method": "mobile_money", "provider": "paystack"})
        denied_branch = self.client.get(reverse("reports:payments"), {"branch": str(self.tse_addo.pk)})
        self.client.force_login(self.cashier)
        denied_role = self.client.get(reverse("reports:payments"))

        self.assertEqual(filtered.status_code, status.HTTP_200_OK)
        self.assertEqual(filtered.json()["summary"]["payment_count"], 1)
        self.assertEqual(filtered.json()["summary"]["net_collected"], "100.00")
        self.assertEqual(denied_branch.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(denied_role.status_code, status.HTTP_403_FORBIDDEN)

    def test_payment_corrections_are_net_zero_and_use_correction_date(self):
        report_day = timezone.localdate() - timezone.timedelta(days=5)
        occurred = timezone.make_aware(datetime.combine(report_day, time(12, 0)))
        original = occurred - timezone.timedelta(days=10)

        refunded_online = Payment.objects.create(
            branch=self.makola, customer=self.customer, order=self.order,
            provider="paystack", method="card", status=Payment.Status.REFUNDED,
            amount="40.00", paid_at=original,
        )
        Payment.objects.filter(pk=refunded_online.pk).update(created_at=original, updated_at=occurred)
        pending = Payment.objects.create(
            branch=self.makola, customer=self.customer, order=self.order,
            method="bank_transfer", status=Payment.Status.PENDING, amount="20.00",
        )
        failed = Payment.objects.create(
            branch=self.makola, customer=self.customer, order=self.order,
            method="card", status=Payment.Status.FAILED, amount="10.00",
        )
        cancelled = Payment.objects.create(
            branch=self.makola, customer=self.customer, order=self.order,
            method="cash", status=Payment.Status.CANCELLED, amount="5.00",
        )
        Payment.objects.filter(pk__in=(pending.pk, failed.pk)).update(created_at=occurred, updated_at=occurred)
        Payment.objects.filter(pk=cancelled.pk).update(created_at=original, updated_at=occurred)

        refunded_sale = POSSale.objects.create(
            branch=self.makola, cashier=self.cashier, status=POSSale.Status.REFUNDED,
            payment_status="refunded", total_amount="30.00", item_count=1,
            completed_at=original,
        )
        refunded_pos = POSPaymentEntry.objects.create(
            sale=refunded_sale, method="cash", amount="30.00", status="refunded",
        )
        POSPaymentEntry.objects.filter(pk=refunded_pos.pk).update(created_at=original, updated_at=occurred)

        self.client.force_login(self.manager)
        response = self.client.get(reverse("reports:payments"), {
            "date_from": report_day.isoformat(), "date_to": report_day.isoformat(),
        })

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.json()["summary"], {
            "payment_count": 5, "successful_count": 0,
            "successful_amount": "70.00", "pending_count": 1,
            "failed_count": 2, "refunded_count": 2,
            "refunded_amount": "70.00", "net_collected": "0.00",
        })
        methods = {row["method"]: row for row in response.json()["by_method"]}
        self.assertEqual(methods["card"]["net_collected"], "0.00")
        self.assertEqual(methods["cash"]["net_collected"], "0.00")
        self.assertEqual(response.json()["daily"], [{
            "date": report_day.isoformat(), "online": "0.00", "pos": "0.00",
            "total": "0.00", "count": 5,
        }])

    def test_sales_date_edges_are_inclusive_and_non_revenue_states_are_excluded(self):
        report_day = timezone.localdate() - timezone.timedelta(days=8)
        start = timezone.make_aware(datetime.combine(report_day, time.min))
        end = timezone.make_aware(datetime.combine(report_day, time.max))

        def order(amount, status_value, payment_status="paid", paid_at=start):
            return Order.objects.create(
                branch=self.makola, customer=self.customer, status=status_value,
                payment_status=payment_status, subtotal=amount, total_amount=amount,
                recipient_name=self.customer.full_name, paid_at=paid_at,
            )

        order("25.00", Order.Status.PAID, paid_at=start)
        order("35.00", Order.Status.DELIVERED, paid_at=end)
        order("100.00", Order.Status.CANCELLED)
        order("100.00", Order.Status.REFUNDED)
        order("100.00", Order.Status.AWAITING_PAYMENT, payment_status="pending")
        order("999.00", Order.Status.PAID, paid_at=start - timezone.timedelta(microseconds=1))

        POSSale.objects.create(branch=self.makola, cashier=self.cashier, status=POSSale.Status.COMPLETED, payment_status="paid", total_amount="15.00", item_count=1, completed_at=start)
        POSSale.objects.create(branch=self.makola, cashier=self.cashier, status=POSSale.Status.COMPLETED, payment_status="paid", total_amount="20.00", item_count=1, completed_at=end)
        POSSale.objects.create(branch=self.makola, cashier=self.cashier, status=POSSale.Status.REFUNDED, payment_status="refunded", total_amount="100.00", item_count=1, completed_at=start)
        POSSale.objects.create(branch=self.makola, cashier=self.cashier, status=POSSale.Status.VOIDED, payment_status="voided", total_amount="100.00", item_count=1, completed_at=start)

        self.client.force_login(self.manager)
        response = self.client.get(reverse("reports:sales"), {
            "date_from": report_day.isoformat(), "date_to": report_day.isoformat(),
        })

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.json()["summary"]["total_revenue"], "95.00")
        self.assertEqual(response.json()["summary"]["transaction_count"], 4)
        self.assertEqual({row["status"] for row in response.json()["transactions"]}, {Order.Status.PAID, Order.Status.DELIVERED, POSSale.Status.COMPLETED})

    def test_cancelled_and_pending_bookings_count_as_demand_but_not_service_revenue(self):
        report_day = timezone.localdate() - timezone.timedelta(days=3)
        appointment_time = timezone.make_aware(datetime.combine(report_day, time(10)))
        cancelled = Booking.objects.create(
            branch=self.makola, customer=self.customer, status=Booking.Status.CANCELLED,
            source=Booking.Source.WEBSITE, preferred_start=appointment_time,
            total_amount="60.00", total_duration_minutes=60, payment_status="paid",
        )
        pending = Booking.objects.create(
            branch=self.makola, customer=self.customer, status=Booking.Status.PENDING,
            source=Booking.Source.WEBSITE, preferred_start=appointment_time,
            total_amount="70.00", total_duration_minutes=60, payment_status="pending",
        )
        BookingServiceItem.objects.create(booking=cancelled, service=self.service, service_name=self.service.name, unit_price="60.00", duration_minutes=60)
        BookingServiceItem.objects.create(booking=pending, service=self.service, service_name=self.service.name, unit_price="70.00", duration_minutes=60)

        self.client.force_login(self.manager)
        filters = {"date_from": report_day.isoformat(), "date_to": report_day.isoformat()}
        bookings = self.client.get(reverse("reports:bookings"), filters)
        services = self.client.get(reverse("reports:services"), filters)

        self.assertEqual(bookings.json()["summary"]["booking_count"], 2)
        self.assertEqual(bookings.json()["summary"]["cancelled_count"], 1)
        self.assertEqual(bookings.json()["summary"]["booked_value"], "130.00")
        self.assertEqual(services.json()["summary"]["service_count"], 2)
        self.assertEqual(services.json()["summary"]["revenue"], "0.00")

    def test_branches_report_combines_cross_operational_metrics(self):
        self.client.force_login(self.manager)
        response = self.client.get(reverse("reports:branches"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.json()["summary"], {
            "branch_count": 1, "total_sales": "150.00", "booking_count": 2,
            "booking_value": "300.00", "payments_collected": "150.00",
            "product_revenue": "150.00", "product_gross_profit": "90.00",
            "service_revenue": "280.00", "estimated_operating_result": "370.00",
            "stock_available": 4,
        })
        branch = response.json()["performance"][0]
        self.assertEqual(branch["branch_id"], str(self.makola.pk))
        self.assertEqual(branch["completed_bookings"], 1)
        self.assertEqual(branch["low_stock_count"], 1)
        self.assertEqual(branch["sales_share_percent"], "100.00")
        self.assertEqual(branch["product_gross_profit"], "90.00")
        self.assertEqual(branch["estimated_operating_result"], "370.00")
        self.assertEqual(branch["cancellation_rate"], "0.00")
        self.assertEqual(branch["no_show_rate"], "0.00")

    def test_owner_can_compare_all_permitted_branches(self):
        self.client.force_login(self.owner)
        response = self.client.get(reverse("reports:branches"))

        self.assertEqual(response.status_code, status.HTTP_200_OK)
        self.assertEqual(response.json()["summary"]["branch_count"], 2)
        self.assertEqual(
            {row["branch_name"] for row in response.json()["performance"]},
            {"Report Makola", "Report Tse Addo"},
        )

    def test_report_summaries_reconcile_to_independent_raw_transaction_totals(self):
        """Control totals come from source rows, independently of report code."""
        self.client.force_login(self.manager)

        paid_orders = Order.objects.filter(branch=self.makola, payment_status="paid").exclude(
            status__in=(Order.Status.CANCELLED, Order.Status.RETURNED, Order.Status.REFUNDED),
        )
        completed_pos = POSSale.objects.filter(branch=self.makola, status=POSSale.Status.COMPLETED)
        raw_online_revenue = sum((row.total_amount for row in paid_orders), Decimal("0.00"))
        raw_pos_revenue = sum((row.total_amount for row in completed_pos), Decimal("0.00"))
        sales = self.client.get(reverse("reports:sales")).json()
        self.assertEqual(Decimal(sales["summary"]["online_revenue"]), raw_online_revenue)
        self.assertEqual(Decimal(sales["summary"]["pos_revenue"]), raw_pos_revenue)
        self.assertEqual(Decimal(sales["summary"]["total_revenue"]), sum((Decimal(row["amount"]) for row in sales["transactions"]), Decimal("0.00")))

        raw_bookings = list(Booking.objects.filter(branch=self.makola))
        bookings = self.client.get(reverse("reports:bookings")).json()
        self.assertEqual(bookings["summary"]["booking_count"], len(raw_bookings))
        self.assertEqual(Decimal(bookings["summary"]["booked_value"]), sum((row.total_amount for row in raw_bookings), Decimal("0.00")))
        self.assertEqual(Decimal(bookings["summary"]["booked_value"]), sum((Decimal(row["amount"]) for row in bookings["bookings"]), Decimal("0.00")))

        raw_order_items = OrderItem.objects.filter(order__in=paid_orders)
        raw_pos_product_lines = POSSaleLine.objects.filter(sale__in=completed_pos, item_type=POSSaleLine.ItemType.PRODUCT)
        raw_product_units = sum(row.quantity for row in raw_order_items) + sum(row.quantity for row in raw_pos_product_lines)
        raw_product_revenue = sum((row.line_total for row in raw_order_items), Decimal("0.00")) + sum((row.line_total for row in raw_pos_product_lines), Decimal("0.00"))
        raw_product_cost = sum((row.line_cost for row in raw_order_items), Decimal("0.00")) + sum((row.line_cost for row in raw_pos_product_lines), Decimal("0.00"))
        products = self.client.get(reverse("reports:products")).json()
        self.assertEqual(products["summary"]["units_sold"], raw_product_units)
        self.assertEqual(Decimal(products["summary"]["revenue"]), raw_product_revenue)
        self.assertEqual(Decimal(products["summary"]["cost_of_goods"]), raw_product_cost)
        self.assertEqual(Decimal(products["summary"]["gross_profit"]), raw_product_revenue - raw_product_cost)
        self.assertEqual(Decimal(products["summary"]["revenue"]), sum((Decimal(row["revenue"]) for row in products["products"]), Decimal("0.00")))

        raw_booking_service_items = BookingServiceItem.objects.filter(
            booking__branch=self.makola, booking__payment_status="paid",
        ).exclude(booking__status__in=(Booking.Status.CANCELLED, Booking.Status.REJECTED))
        raw_pos_service_lines = POSSaleLine.objects.filter(sale__in=completed_pos, item_type=POSSaleLine.ItemType.SERVICE)
        raw_service_revenue = sum((row.unit_price for row in raw_booking_service_items), Decimal("0.00")) + sum((row.line_total for row in raw_pos_service_lines), Decimal("0.00"))
        services = self.client.get(reverse("reports:services")).json()
        self.assertEqual(Decimal(services["summary"]["revenue"]), raw_service_revenue)
        self.assertEqual(Decimal(services["summary"]["revenue"]), sum((Decimal(row["revenue"]) for row in services["performance"]), Decimal("0.00")))

        raw_inventory = list(BranchInventory.objects.filter(branch=self.makola))
        inventory = self.client.get(reverse("reports:inventory")).json()
        self.assertEqual(inventory["summary"]["quantity_on_hand"], sum(row.quantity_on_hand for row in raw_inventory))
        self.assertEqual(inventory["summary"]["quantity_reserved"], sum(row.quantity_reserved for row in raw_inventory))
        self.assertEqual(inventory["summary"]["quantity_available"], sum(row.quantity_available for row in raw_inventory))
        self.assertEqual(inventory["summary"]["on_hand_change"], sum(StockMovement.objects.filter(inventory__branch=self.makola).values_list("quantity_on_hand_change", flat=True)))

        raw_online_payments = Payment.objects.filter(branch=self.makola)
        raw_pos_payments = POSPaymentEntry.objects.filter(sale__branch=self.makola)
        raw_gross = sum((row.amount for row in raw_online_payments if row.status in (Payment.Status.SUCCEEDED, Payment.Status.REFUNDED)), Decimal("0.00")) + sum((row.amount for row in raw_pos_payments if row.status in ("succeeded", "refunded")), Decimal("0.00"))
        raw_refunds = sum((row.amount for row in raw_online_payments if row.status == Payment.Status.REFUNDED), Decimal("0.00")) + sum((row.amount for row in raw_pos_payments if row.status == "refunded"), Decimal("0.00"))
        payments = self.client.get(reverse("reports:payments")).json()
        self.assertEqual(Decimal(payments["summary"]["successful_amount"]), raw_gross)
        self.assertEqual(Decimal(payments["summary"]["refunded_amount"]), raw_refunds)
        self.assertEqual(Decimal(payments["summary"]["net_collected"]), raw_gross - raw_refunds)
        self.assertEqual(Decimal(payments["summary"]["net_collected"]), sum((Decimal(row["net_collected"]) for row in payments["by_method"]), Decimal("0.00")))

        branches = self.client.get(reverse("reports:branches")).json()
        self.assertEqual(Decimal(branches["summary"]["total_sales"]), sum((Decimal(row["total_sales"]) for row in branches["performance"]), Decimal("0.00")))
        self.assertEqual(Decimal(branches["summary"]["product_revenue"]), sum((Decimal(row["product_revenue"]) for row in branches["performance"]), Decimal("0.00")))
        self.assertEqual(Decimal(branches["summary"]["service_revenue"]), sum((Decimal(row["service_revenue"]) for row in branches["performance"]), Decimal("0.00")))
        self.assertEqual(
            Decimal(branches["summary"]["estimated_operating_result"]),
            Decimal(branches["summary"]["product_gross_profit"]) + Decimal(branches["summary"]["service_revenue"]),
        )

    def test_branch_exports_label_estimated_result_and_include_cost_limitation(self):
        self.client.force_login(self.manager)
        export_url = reverse("reports:export", kwargs={"report_name": "branches"})
        csv_response = self.client.get(export_url, {"file_format": "csv"})
        workbook_response = self.client.get(export_url, {"file_format": "xlsx"})

        csv_text = csv_response.content.decode("utf-8-sig")
        self.assertIn("Estimated Operating Result", csv_text)
        self.assertIn("Important limitation", csv_text)
        self.assertIn("not net profit", csv_text)
        workbook = load_workbook(BytesIO(workbook_response.content), read_only=True)
        self.assertEqual(workbook["Summary"]["A8"].value, "Important limitation")
        self.assertIn("not net profit", workbook["Summary"]["B8"].value)

    def test_all_management_reports_export_as_branded_pdf(self):
        self.client.force_login(self.manager)
        for report_name in ("sales", "bookings", "products", "services", "inventory", "payments", "branches"):
            with self.subTest(report=report_name):
                response = self.client.get(
                    reverse("reports:export", kwargs={"report_name": report_name}),
                    {"file_format": "pdf"},
                )
                self.assertEqual(response.status_code, status.HTTP_200_OK, response.content)
                self.assertEqual(response["Content-Type"], "application/pdf")
                self.assertTrue(response.content.startswith(b"%PDF"))
                self.assertIn(f"golden-touch-{report_name}-report.pdf", response["Content-Disposition"])

    def test_report_exports_real_excel_and_utf8_csv(self):
        self.client.force_login(self.manager)
        export_url = reverse("reports:export", kwargs={"report_name": "sales"})
        excel = self.client.get(export_url, {"file_format": "xlsx"})
        csv_response = self.client.get(export_url, {"file_format": "csv"})

        self.assertEqual(excel.status_code, status.HTTP_200_OK)
        workbook = load_workbook(BytesIO(excel.content), read_only=True)
        self.assertEqual(workbook["Summary"]["A1"].value, "Online Orders and POS Sales Report")
        self.assertIn("Transactions", workbook.sheetnames)
        self.assertEqual(csv_response.status_code, status.HTTP_200_OK)
        csv_text = csv_response.content.decode("utf-8-sig")
        self.assertIn("Online Orders and POS Sales Report", csv_text)
        self.assertIn("Transactions", csv_text)

    def test_report_export_permissions_and_format_are_enforced(self):
        export_url = reverse("reports:export", kwargs={"report_name": "sales"})
        self.client.force_login(self.cashier)
        denied = self.client.get(export_url, {"file_format": "pdf"})
        self.client.force_login(self.manager)
        invalid = self.client.get(export_url, {"file_format": "zip"})

        self.assertEqual(denied.status_code, status.HTTP_403_FORBIDDEN)
        self.assertEqual(invalid.status_code, status.HTTP_400_BAD_REQUEST)

    def test_branches_report_filters_and_permissions_are_enforced(self):
        self.client.force_login(self.manager)
        filtered = self.client.get(reverse("reports:branches"), {"branch": str(self.makola.pk), "sort": "name"})
        denied_branch = self.client.get(reverse("reports:branches"), {"branch": str(self.tse_addo.pk)})
        self.client.force_login(self.cashier)
        denied_role = self.client.get(reverse("reports:branches"))

        self.assertEqual(filtered.status_code, status.HTTP_200_OK)
        self.assertEqual(filtered.json()["summary"]["branch_count"], 1)
        self.assertEqual(denied_branch.status_code, status.HTTP_400_BAD_REQUEST)
        self.assertEqual(denied_role.status_code, status.HTTP_403_FORBIDDEN)
